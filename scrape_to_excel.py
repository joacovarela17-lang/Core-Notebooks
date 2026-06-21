import json
import urllib.request
import os

def get_blue_rate():
    try:
        import ssl
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        
        url = 'https://dolarapi.com/v1/dolares/blue'
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, context=ctx, timeout=5) as response:
            data = json.loads(response.read().decode('utf-8'))
            if data and 'venta' in data:
                return float(data['venta'])
    except Exception as e:
        print(f"Error fetching live blue rate: {e}. Using fallback 1250.")
    return 1250.0

def generate_excel():
    blue_rate = get_blue_rate()
    catalog_path = 'catalog.json'
    output_path = 'precios_venta.xlsx'
    
    if not os.path.exists(catalog_path):
        print(f"Error: {catalog_path} not found.")
        return
        
    with open(catalog_path, 'r', encoding='utf-8') as f:
        products = json.load(f)
        
    # Filter out products with cost <= $700
    filtered_products = [p for p in products if float(p.get('price_usd', 0)) > 700.0]
    
    headers = [
        "ID", "Marca", "Modelo", "Categoría", "Tipo", 
        "Costo USD", "Precio Venta USD", "Precio Venta ARS", "Ganancia USD"
    ]
    
    rows = []
    for p in filtered_products:
        cost = float(p.get('price_usd', 0))
        selling_usd = 1.20 * cost + 140
        selling_ars = selling_usd * blue_rate
        profit_usd = selling_usd - cost
        
        rows.append([
            p.get('id', ''),
            p.get('brand', ''),
            p.get('name', ''),
            p.get('category', ''),
            p.get('type', ''),
            cost,
            round(selling_usd, 2),
            round(selling_ars, 2),
            round(profit_usd, 2)
        ])
        
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Precios de Venta"
        
        # Enable grid lines
        ws.views.sheetView[0].showGridLines = True
        
        # Styles
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0071E3", end_color="0071E3", fill_type="solid") # Apple Blue
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        data_font = Font(name="Segoe UI", size=10)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        
        thin_side = Side(border_style="thin", color="D3D3D3")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Write headers
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            
        # Write data
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                
                # Alignments and number formats
                if c_idx in [1, 4, 5]:  # ID, Category, Type
                    cell.alignment = center_align
                elif c_idx in [2, 3]:   # Brand, Model
                    cell.alignment = left_align
                else:                   # Cost, Prices, Profit
                    cell.alignment = right_align
                    if c_idx == 8:
                        cell.number_format = '"$"#,##0.00'
                    else:
                        cell.number_format = '"U$S "#,##0.00'
                        
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            # Add padding
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
            
        wb.save(output_path)
        print(f"Spreadsheet generated successfully at {output_path} with {len(rows)} products (Blue rate: {blue_rate}).")
        
    except ImportError:
        # Fallback to a clean semicolon separated CSV that Excel opens directly
        print("openpyxl not installed. Generating Excel-compatible semicolon separated CSV file.")
        import csv
        with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerow(headers)
            for r in rows:
                # Format numbers for better CSV-Excel compatibility
                r[5] = str(r[5]).replace('.', ',')
                r[6] = str(r[6]).replace('.', ',')
                r[7] = str(r[7]).replace('.', ',')
                r[8] = str(r[8]).replace('.', ',')
                writer.writerow(r)
        print(f"Excel-compatible CSV generated successfully at {output_path} with {len(rows)} products.")

if __name__ == '__main__':
    generate_excel()
